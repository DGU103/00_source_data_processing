import os
import openpyxl
from openpyxl.styles import PatternFill
import shutil

def compare_excel_files(folder):
    # Get list of files in the folder
    files = [f for f in os.listdir(folder) if f.endswith('.xlsx')]

    # Initialize log file
    log_file = open('comparison_log.txt', 'w')

    # Pick the first file in the folder for the copy and also consider it as a source file for comparison
    result_file = os.path.join(folder, 'comparison_result.xlsx')
    shutil.copy(os.path.join(folder, files[1]), result_file)

    # Load the result workbook
    result_wb = openpyxl.load_workbook(result_file, data_only=True)

    file1 = os.path.join(folder, files[0])
    file2 = os.path.join(folder, files[1])

    # Load workbooks
    wb1 = openpyxl.load_workbook(file1, data_only=True)
    wb2 = openpyxl.load_workbook(file2, data_only=True)

    # Compare sheets
    for sheet_name in wb1.sheetnames:
        if sheet_name in wb2.sheetnames:
            ws1 = wb1[sheet_name]
            ws2 = wb2[sheet_name]
            result_ws = result_wb[sheet_name]

            # Compare cells
            for row in range(1, ws1.max_row + 1):
                for col in range(1, ws1.max_column + 1):
                    cell1 = ws1.cell(row=row, column=col)
                    cell2 = ws2.cell(row=row, column=col)

                    result_cell = result_ws.cell(row=row, column=col)

                    if not isinstance(cell2, openpyxl.cell.cell.MergedCell):
                        value1 = cell1.value
                        value2 = cell2.value

                        # Try to convert values to numerical before comparison
                        try:
                            value1 = float(value1)
                        except (ValueError, TypeError):
                            pass

                        try:
                            value2 = float(value2)
                        except (ValueError, TypeError):
                            pass

                        if value1 is not None and value2 is not None:
                            if isinstance(value1, (int, float)) and isinstance(value2, (int, float)):
                                # Compare numeric values with tolerance of 0.5%
                                if abs(value1 - value2) > 0.005 * abs(value1):
                                    if not isinstance(result_cell, openpyxl.cell.cell.MergedCell):
                                        result_cell.value = f"{value2} (diff: {value1})"
                                        result_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                                else:
                                    result_cell.value = value2
                            else:
                                # Compare other types of values directly
                                if value1 != value2:
                                    if not isinstance(result_cell, openpyxl.cell.cell.MergedCell):
                                        result_cell.value = f"{value2} (diff: {value1})"
                                        result_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                                else:
                                    result_cell.value = value2
                        elif value1 is not None or value2 is not None:
                            # Handle case where one cell is None and the other is not
                            if not isinstance(result_cell, openpyxl.cell.cell.MergedCell):
                                result_cell.value = f"{value2} (diff: {value1})"
                                result_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        else:
                            result_cell.value = value2
        else:
            log_file.write(f"Sheet '{sheet_name}' is missing in file '{files}'.\n")

    # Save the result workbook
    result_wb.save(result_file)
    log_file.close()


# Example usage
folder = 'C:/Users/mch107/Downloads/Doc_Comparison/'

compare_excel_files(folder)
